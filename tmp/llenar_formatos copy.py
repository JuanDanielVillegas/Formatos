import months 
import pyodbc
import os
from openpyxl import load_workbook

from celdas import tags, tags_sima
from funciones_aux import *
from acceso import *

def set_data_to_template(wb, ws, emp, sima, template):

    if sima != 1:
        for tag in tags:
            header = tags[tag]
            content = emp[tags[tag]]
            underscore = '______'
            cell = ws[tag]
            
            cell.value = header + ':' + underscore + content + underscore
            #print(ws[tag].value)
    #else:
      #for tag in tags_sima:
            #header = tags_sima[tag]
            #content = 'pendiente' 
            #underscore = '______'
            #cell = ws[tag]
            
            
            #print(ws[tag].value)
    save_file(wb, emp['Nombre'], template)
  

def fill_templates(emp):

    templates = [
        'CISCO',
        'DII',
        'EFA',
        'EVAC',
        'SIMA'
    ]

    for t in templates: 
        template_name = t + '.xlsx'
        path = './plantillas_formatos/' + template_name

        wb = load_workbook(path)
        ws = wb['F1']

        if t != 'SIMA':
            set_data_to_template(wb, ws, emp, 0, t)
        else:
            set_data_to_template(wb, ws, emp, 1, t)


def save_file(wb, name, template):

    
    file_name = template + '_' + name + '.xlsx'

    path = './plantillas_formatos/Formatos/' + name + '/' + file_name

    wb.save(path)
    


def find_emp(): 
    nom = input('Nombre: ')
    apat = input('Apellido paterno: ')
    amat = input('Apellido materno: ')

    try:
        connection = pyodbc.connect(conn_str)
        cursor = connection.cursor()

        cursor.execute(query, ('%' + nom + '%', '%' + apat + '%', '%' +amat + '%'))

        rows = cursor.fetchall()

        emp = {
            'Fecha de solicitud' : get_date(),
            'Departamento' : rows[0].Departamento,
            'Nombre' : rows[0].Nombre,
            'Puesto' : rows[0].Puesto,
            'RFC(con homoclave)' : rows[0].RFC,
            'CURP' : rows[0].Curp,
            'Número o Clave de Empleado' : rows[0].Numero,
            'Dirección de correo electrónico' : rows[0].correo.replace('BUZON.', ''),
            'Teléfono': '(614) 429-33-00 Ext. ' + rows[0].ext,
            'Dirección IP': '10.18.101.82',
            'Nombre y Firma' : rows[0].titulo.upper() + ' ' + rows[0].Nombre  
        }

        return emp

    except pyodbc.Error as e:
        print(f'Error de conexión: {e}')

    finally:
        if connection:
            connection.close()


def select_emp():
    emp = find_emp()
    os.system('cls')
    print('Numero de empleado \t Nombre')
    print(emp['Número o Clave de Empleado'],'\t\t', emp['Nombre'])

    opt = input('Seleccionar (s/n): ')

    if opt == 's':
        path = './plantillas_formatos/Formatos/' + emp['Nombre']
        os.mkdir(path)
        fill_templates(emp)
    else:
        select_emp()


select_emp()

    

