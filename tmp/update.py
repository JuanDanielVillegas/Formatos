from openpyxl import load_workbook
from datetime import datetime
import months 
import pyodbc

def connect():
    server_name = 'INTRAFIGO_SVR'
    database_name = 'fiscalizacion'
    username = 'sa'
    password = 'Systemas07'
    conn_str = f'DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID={username};PWD={password}'

    nom = input('Nombre: ')
    apat = input('Apellido paterno: ')
    amat = input('Apellido materno: ')

    try:
        connection = pyodbc.connect(conn_str)
        cursor = connection.cursor()

        sql_query = """
            SELECT
                emp.str_empleado AS 'Numero',
                titulo.str_titulo AS titulo,
                CONCAT(emp.str_nombre, ' ', emp.str_apat, ' ', emp.str_amat) AS 'Nombre', 
                emp.str_curp AS Curp,
                emp.str_rfc AS RFC,
                puesto.str_puesto AS 'Puesto',
                depto.str_depto AS Departamento,
                emp.str_email AS correo,
                CONCAT(sup.str_nombre, ' ', sup.str_apat, ' ', sup.str_amat) AS 'Nombre del superior'
            FROM 
                dd_empleados AS emp
                INNER JOIN dd_empleados AS sup ON emp.c_superior = sup.c_empleado
                INNER JOIN mm_departamento AS depto ON depto.c_depto = emp.c_depto
                INNER JOIN mm_ciudad AS ciudad ON ciudad.c_ciudad = emp.c_ciudad
                INNER JOIN mm_puesto AS puesto ON puesto.c_puesto = emp.c_puesto
                INNER JOIN mm_titulo AS titulo ON titulo.c_titulo = emp.c_titulo
            WHERE 
                emp.c_ciudad = 1 AND emp.bol_vige = 1 
                AND emp.str_nombre  COLLATE Latin1_General_100_CI_AI LIKE ? COLLATE Latin1_General_100_CI_AI
                AND emp.str_apat  COLLATE Latin1_General_100_CI_AI LIKE ? COLLATE Latin1_General_100_CI_AI
                AND emp.str_amat  COLLATE Latin1_General_100_CI_AI LIKE ? COLLATE Latin1_General_100_CI_AI;
        """
        cursor.execute(sql_query, ('%' + nom + '%', '%' + apat + '%', '%' +amat + '%'))

        rows = cursor.fetchall()

        
        depto = rows[0].Departamento
        nombre = rows[0].Nombre
        puesto = rows[0].Puesto
        rfc = rows[0].RFC
        curp = rows[0].Curp
        numero_emp = rows[0].Numero
        correo = rows[0].correo.replace('BUZON.', '')
        titulo = rows[0].titulo


        print(depto+'| '+nombre+'| '+puesto+'| '+rfc+'| '+curp+'| '+ numero_emp+'| '+correo+'| '+titulo.upper())


        #for row in rows:
        #    print(row)

    except pyodbc.Error as e:
        print(f'Error de conexión: {e}')

    finally:
        if connection:
            connection.close()



def get_month(n):
    months = [
        'Enero',
        'Febrero',
        'Marzo',
        'Abril',
        'Mayo',
        'Junio',
        'Julio',
        'Agosto',
        'Septiembre',
        'Octubre',
        'Noviembre',
        'Diciembre'
    ]

    return months[n-1]

#def create_form(f):

def set_data(wb, ws, tags):

    now = datetime.now()
    date = str(now.day) + ' de ' + get_month(now.month) + ' ' + str(now.year)

    for tag in tags:
        header = tags[tag]
        content = 'in progress' 
        underscore = '______'
        cell = ws[tag]
        print(ws[tag].value)

    print(date)


    #cell.value = header + ':' + underscore + content + underscore


    emp = 'Ejemplo'
    file_name = 
    #wb.save('CISCO_TEMPLATE.xlsx')


wb = load_workbook("./plantillas_formatos/CISCO.xlsx")
ws = wb['F1']

#SIMA 
tags = {
    'A8': 'Fecha de solicitud',
    'A14': 'Departamento',
    'A18': 'Nombre',
    'A19': 'Puesto',
    'A20': 'RFC(con homoclave)',
    'C20': 'CURP',
    'A21': 'Número o Clave de Empleado',
    'A22': 'Dirección de correo electrónico',
    'A34': 'Dirección IP',
    'A81': 'Nombre y Firma'
}

templates = [
    'CISCO',
    'DII',
    'EFA',
    'EVAC',
    'SIMA'
]


set_data(wb, ws, tags)

