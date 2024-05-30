from tkinter import *
import customtkinter 

import xlsxwriter
from cells import *
from openpyxl import load_workbook
from functions import *


def append_apps(checkboxes):
    apps = [
        'DII',
        'CISCO',
        'EFA',
        'EVAC'
    ]


    checkboxes_values = {
        'SUIEFI': checkboxes[0].get(),
        'DYP': checkboxes[1].get(),
        'CU_WEB': checkboxes[2].get(),
        'COGNOS_WEB': checkboxes[3].get(),
        'CRP': checkboxes[4].get(),
        'SIMA_PLAN': checkboxes[5].get(),
        'SAT_CLOUD': checkboxes[6].get(),
        'VUCEM': checkboxes[7].get(),
        'SIMA_SEG': checkboxes[8].get()
    }


    for a in checkboxes_values:
        if(checkboxes_values[a] == 1):
            apps.append(a)


    return apps

def fill_templates(emp_name, checked, suiefi, cedula, st, s, checkboxes, folder_date):

    template_path = './plantillas_formatos/PLANTILLA.xlsx'
    ws_template = read_template(template_path)


    apps = append_apps(checkboxes)
 


    for app_name in apps: 
        write_form(ws_template, app_name, emp_name, checked, suiefi, cedula, st, s, folder_date)


def read_template(path):
    wb = load_workbook(path)
    ws = wb['F1']
    return ws


def write_form(workbook_template, app_name, emp_name, checked, suiefi, cedula, st, s, folder_date):
    empty_cells = []
    file_name = './Formatos/'+ folder_date + '/' + emp_name.strip() + '/' + app_name + '_'+emp_name.strip() +'.xlsx' ## --- Here
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()
    worksheet.set_column(0, 0, 26)
    worksheet.set_column('B:G', 14)

    size_offset = {"x_offset": 30, "y_offset": 5, "x_scale": 1.3, "y_scale": 1.3}
    worksheet.insert_image('A1', 'shcp.png' , size_offset)
    worksheet.insert_image('F1', 'sat.png' , size_offset)

    underline = workbook.add_format({'underline':True, 'font_name': 'Arial', 'font_size': 8})
    template_tags = fill_template_tags()
    s = '                         '
    empty_string = s + s + s + s + s + s + s

    #shrink rows
    for r in rows:
        if r in [42, 48, 58]:
            worksheet.set_row(r-1, 6)
        else: 
            worksheet.set_row(r-1, 3)

    if app_name == 'SIMA_SEG' or app_name == 'SIMA_PLAN':
        worksheet.set_row(24, 15)
        worksheet.set_row(25, 22)

    #set cell borders
    for b in borders:
        worksheet.write_column(b, ' ', workbook.add_format({'border': 2}))
    
    content = 'X'
    worksheet.write_column(checked, content, workbook.add_format({'border': 2, 'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))

    if app_name == 'SUIEFI':
        if suiefi[0] != False:
            worksheet.write_column(suiefi[0], content, workbook.add_format({'border': 2, 'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))
        if suiefi[1] != 0:
            worksheet.write_column(suiefi[1], content, workbook.add_format({'border': 2, 'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))
        if suiefi[2] != 0:
            worksheet.write_column(suiefi[2], content, workbook.add_format({'border': 2, 'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))
        if suiefi[3] != 0:
            worksheet.write_column(suiefi[3], content, workbook.add_format({'border': 2, 'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))



    for t in template_tags:

        if not workbook_template[t].value is None:
            if t in cells:
            
                content = '  ' + str(cells[t]) + '            ' ## ----- Here
                worksheet.write_rich_string(t, workbook.add_format({'font_name': 'Arial', 'font_size': 8}), workbook_template[t].value, underline, content)

                if t == 'A8': 
                    worksheet.write_rich_string(t, workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'bold': True}), workbook_template[t].value, workbook.add_format({'underline':True, 'font_name': 'Arial', 'font_size': 8, 'bold': True}), content)
                if t == 'A75' or t == 'A81':
                    content = '  ' + cells[t] + empty_string
                    worksheet.write_rich_string(t, workbook.add_format({'font_name': 'Arial', 'font_size': 8}), workbook_template[t].value, underline, content)
            
            else: 

                if t == 'A36' or t == 'A37':
                    content = ' ' + app[app_name]['name'] + '            '
                    worksheet.write_rich_string(t, cell_format, workbook_template[t].value, underline, content)
                    content = ' ' + app[app_name]['profile'] + '            '
                    worksheet.write_rich_string('A37', cell_format, workbook_template[t].value, underline, content)
                    continue

                content = ' ' + str(workbook_template[t].value) + '            '
                style = {'font_name': 'Arial', 'font_size': 8}

                cell_format = workbook.add_format(style)
                worksheet.write(t, content, cell_format)
                
                if t == 'A4': 
                    worksheet.merge_range("A4:G4", content, workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bg_color':'#BFBFBF'}))
                
                if t in center_bold: 
                    r = t + ':' + t.replace('A', 'G')
                    content = str(workbook_template[t].value) + '            '
                    worksheet.merge_range(r, content, workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'align':'center', 'bold':True}))
                
                if t in bold: 
                    content = str(workbook_template[t].value) + '            '
                    worksheet.write(t, content, workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'bold':True}))
                
                if t in align_right: 
                    content = str(workbook_template[t].value)
                    worksheet.write(t, content, workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'align':'right'}))
                
                if t == 'A56':

                    worksheet.merge_range('A56:A57', str(workbook_template[t].value), workbook.add_format({'font_name': 'Arial', 'font_size': 8, 'text_wrap':True}))


                if t == 'A62':
                    content = app[app_name]['explication'] + '            '
                    worksheet.merge_range('A62:G64', content, workbook.add_format({'font_name': 'Arial', 'font_size': 10, 'underline':True, 'text_wrap':True}))
        else:
            
            if (app_name == 'SIMA_PLAN' or app_name == 'SIMA_SEG') and t == 'A25':
                content = '  ' + cells[t] + '            '
                worksheet.write_rich_string(t, workbook.add_format({'font_name': 'Arial', 'font_size': 8}), 'Nombre del Jefe al que Reporta:', underline, content)

            if 'A' in t:
                empty_cells.append(t)

    workbook.close()
